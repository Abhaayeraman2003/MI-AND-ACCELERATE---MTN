"""
Excel storage for the DGW MI & Accelerate app.

The master workbook (submissions.xlsx) lives in YOUR SharePoint. The app writes to it
using a service credential that the OpCo submitters never have — so people entering data
can use the form but CANNOT open, download or edit the master file. Only you (via
SharePoint) can.

Sheets in the master workbook
  * "Submissions" — clean flat table (one row per initiative) incl. date + time stamps
  * "_audit"      — an append-only log: every submit event with UTC timestamp, OpCo, user
  * "_raw"        — hidden JSON of each submission (used to rebuild the dashboard)

Re-submitting the same OpCo + Year + Month overwrites that OpCo's rows (no duplicates),
but the _audit log keeps every event so you have a full history.

Storage backend
  * SharePoint (recommended)  — set st.secrets["sharepoint"] (see secrets.toml.example).
                                Uses Microsoft Graph app-only auth (client credentials).
  * Local file (default)      — ./submissions.xlsx, for running on your own PC.
"""

import io
import os
import json
import datetime

from openpyxl import Workbook, load_workbook

import common

# --------------------------------------------------------------------------- #
LOCAL_PATH = os.environ.get("MASTER_XLSX",
                            os.path.join(os.path.dirname(__file__), "submissions.xlsx"))

FLAT_HEADERS = ["OpCo", "Reporting month", "Submitted by", "Email", "Type", "Section",
                "Initiative", "RAG", "Accelerate %", "Actual", "Estimated",
                "Maturity %", "Current", "Target", "Comment / risk",
                "Submitted (UTC)", "Submitted date", "Submitted time"]

AUDIT_HEADERS = ["Timestamp (UTC)", "Date", "Time", "OpCo", "Reporting month",
                 "Submitted by", "Email", "Initiatives updated", "Action"]


# --------------------------------------------------------------------------- #
#  SharePoint (Microsoft Graph) helpers
# --------------------------------------------------------------------------- #
def _sp_conf():
    """Return the SharePoint config dict from st.secrets, or None if not set."""
    try:
        import streamlit as st
        if "sharepoint" in st.secrets:
            c = dict(st.secrets["sharepoint"])
            need = ("tenant_id", "client_id", "client_secret", "site_hostname",
                    "site_path", "file_path")
            if all(c.get(k) for k in need):
                return c
    except Exception:
        pass
    return None


def _graph_token(c):
    import requests
    url = "https://login.microsoftonline.com/%s/oauth2/v2.0/token" % c["tenant_id"]
    data = {"client_id": c["client_id"], "client_secret": c["client_secret"],
            "scope": "https://graph.microsoft.com/.default", "grant_type": "client_credentials"}
    r = requests.post(url, data=data, timeout=30)
    r.raise_for_status()
    return r.json()["access_token"]


def _graph_site_drive(c, token):
    """Resolve the SharePoint site id and default drive id."""
    import requests
    h = {"Authorization": "Bearer " + token}
    site = requests.get(
        "https://graph.microsoft.com/v1.0/sites/%s:%s" % (c["site_hostname"], c["site_path"]),
        headers=h, timeout=30)
    site.raise_for_status()
    site_id = site.json()["id"]
    drive = requests.get("https://graph.microsoft.com/v1.0/sites/%s/drive" % site_id,
                         headers=h, timeout=30)
    drive.raise_for_status()
    return site_id, drive.json()["id"]


def _sp_download_bytes(c):
    """Download the master workbook from SharePoint, or None if it doesn't exist yet."""
    import requests
    token = _graph_token(c)
    _sid, drive_id = _graph_site_drive(c, token)
    h = {"Authorization": "Bearer " + token}
    url = "https://graph.microsoft.com/v1.0/drives/%s/root:/%s:/content" % (drive_id, c["file_path"])
    r = requests.get(url, headers=h, timeout=60)
    if r.status_code == 404:
        return None
    r.raise_for_status()
    return r.content


def _sp_upload_bytes(c, data):
    """Upload (create/replace) the master workbook to SharePoint."""
    import requests
    token = _graph_token(c)
    _sid, drive_id = _graph_site_drive(c, token)
    h = {"Authorization": "Bearer " + token,
         "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    url = "https://graph.microsoft.com/v1.0/drives/%s/root:/%s:/content" % (drive_id, c["file_path"])
    r = requests.put(url, headers=h, data=data, timeout=120)
    r.raise_for_status()
    return True


# --------------------------------------------------------------------------- #
#  Load / save the workbook bytes (SharePoint or local)
# --------------------------------------------------------------------------- #
def _load_wb():
    c = _sp_conf()
    if c:
        data = _sp_download_bytes(c)
        if data:
            return load_workbook(io.BytesIO(data))
        return _new_wb()
    if os.path.exists(LOCAL_PATH):
        return load_workbook(LOCAL_PATH)
    return _new_wb()


def _save_wb(wb):
    c = _sp_conf()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    if c:
        _sp_upload_bytes(c, buf.read())
    else:
        with open(LOCAL_PATH, "wb") as fh:
            fh.write(buf.read())


def _new_wb():
    wb = Workbook()
    ws = wb.active
    ws.title = "Submissions"
    ws.append(FLAT_HEADERS)
    au = wb.create_sheet("_audit")
    au.append(AUDIT_HEADERS)
    raw = wb.create_sheet("_raw")
    raw.append(["opco", "year", "month", "payload_json"])
    raw.sheet_state = "hidden"
    return wb


# --------------------------------------------------------------------------- #
def _read_raw(wb):
    if "_raw" not in wb.sheetnames:
        return []
    out = []
    for r in wb["_raw"].iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        opco, year, month, pj = (list(r) + [None] * 4)[:4]
        try:
            payload = json.loads(pj) if pj else {}
        except Exception:
            payload = {}
        out.append((opco, str(year), str(month), payload))
    return out


def _style_header(ws):
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    for cc in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=cc)
        cell.font = Font(bold=True, color="0B0B0B")
        cell.fill = PatternFill("solid", fgColor="FFCB05")
    widths = [16, 16, 18, 24, 20, 20, 34, 12, 12, 14, 14, 12, 13, 13, 40, 22, 14, 12]
    for i, w in enumerate(widths[:ws.max_column], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _split_ts(iso):
    """'2026-08-13T09:30:00Z' -> ('2026-08-13T09:30:00Z','2026-08-13','09:30:00')."""
    s = (iso or "").replace("Z", "")
    d = t = ""
    if "T" in s:
        d, t = s.split("T", 1)
        t = t.split(".")[0]
    return iso or "", d, t


import re as _re


def _sheet_title(opco, used):
    """Excel-safe, unique sheet name for an OpCo (<=31 chars, no illegal chars)."""
    name = _re.sub(r"[\\/?*\[\]:]", " ", str(opco or "OpCo")).strip()[:31] or "OpCo"
    base = name
    i = 2
    while name.lower() in used:
        suffix = " (%d)" % i
        name = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(name.lower())
    return name


def _payload_rows(payload):
    """Yield the flat row-values for one submission payload."""
    iso, dpart, tpart = _split_ts(payload.get("submittedAt", ""))
    for rec in common.submissions_to_records([payload]):
        f = rec["fields"]
        ap = act = est = mp = cur = tgt = ""
        if rec["kind"] == "Accelerate":
            p, a, e, _ek, _ak = common.accel_progress(f)
            ap = round(p, 1) if p is not None else ""
            act = a if a is not None else ""
            est = e if e is not None else ""
        if rec["kind"] == "MI":
            p, cu, tg = common.maturity_progress(f)
            mp = round(p, 0) if p is not None else ""
            cur, tgt = cu or "", tg or ""
        rag = common.rag_bucket(f.get("RAG Status", ""))
        yield rag, [rec["opco"], payload.get("reportingMonth", ""), payload.get("submittedBy", ""),
                    payload.get("email", ""), rec["type"], rec.get("section", ""), rec["initiative"],
                    rag, ap, act, est, mp, cur, tgt, f.get("Comment / risk", ""),
                    iso, dpart, tpart]


def _rebuild_flat(wb, raws):
    """Rebuild the workbook with ONE SHEET PER OPCO (plus an 'Overview' summary)."""
    from openpyxl.styles import PatternFill, Font
    RAGFILL = {"Green": "E8F5E9", "Amber": "FFF6E0", "Red": "FDECEC", "Blue": "EAF2FC"}

    # remove any previously-built visible data sheets (keep _audit and _raw)
    for sn in list(wb.sheetnames):
        if sn not in ("_audit", "_raw"):
            del wb[sn]

    # group submissions by OpCo (keep latest period order as inserted)
    by_opco = {}
    for opco, year, month, payload in raws:
        by_opco.setdefault(opco, []).append(payload)

    # ---- Overview sheet (one row per OpCo) ----
    ov = wb.create_sheet("Overview", 0)
    ov.append(["OpCo", "Reporting month", "Submitted by", "Email",
               "Initiatives", "Updated", "Submitted date", "Submitted time"])
    for cc in range(1, 9):
        ov.cell(row=1, column=cc).font = Font(bold=True, color="0B0B0B")
        ov.cell(row=1, column=cc).fill = PatternFill("solid", fgColor="FFCB05")

    used = {"overview"}
    for opco in by_opco:
        payloads = by_opco[opco]
        title = _sheet_title(opco, used)
        ws = wb.create_sheet(title)
        ws.append(FLAT_HEADERS)
        n_init = 0
        for payload in payloads:
            for rag, vals in _payload_rows(payload):
                ws.append(vals)
                n_init += 1
                if rag in RAGFILL:
                    ws.cell(row=ws.max_row, column=8).fill = PatternFill("solid", fgColor=RAGFILL[rag])
        _style_header(ws)

        latest = payloads[-1]
        _iso, dpart, tpart = _split_ts(latest.get("submittedAt", ""))
        ov.append([opco, latest.get("reportingMonth", ""), latest.get("submittedBy", ""),
                   latest.get("email", ""), n_init, latest.get("itemsUpdated", ""),
                   dpart, tpart])

    for w, col in zip([18, 16, 20, 26, 12, 10, 14, 12], "ABCDEFGH"):
        ov.column_dimensions[col].width = w
    ov.freeze_panes = "A2"

    # keep the audit log as the last visible tab for tidiness
    if "_audit" in wb.sheetnames:
        wb.move_sheet("_audit", offset=len(wb.sheetnames))


def _append_audit(wb, payload, action):
    from openpyxl.styles import Font, PatternFill
    if "_audit" not in wb.sheetnames:
        au = wb.create_sheet("_audit")
        au.append(AUDIT_HEADERS)
        for cc in range(1, len(AUDIT_HEADERS) + 1):
            au.cell(row=1, column=cc).font = Font(bold=True, color="0B0B0B")
            au.cell(row=1, column=cc).fill = PatternFill("solid", fgColor="FFCB05")
    au = wb["_audit"]
    iso, dpart, tpart = _split_ts(payload.get("submittedAt", ""))
    au.append([iso, dpart, tpart, payload.get("opco", ""), payload.get("reportingMonth", ""),
               payload.get("submittedBy", ""), payload.get("email", ""),
               payload.get("itemsUpdated", ""), action])


# --------------------------------------------------------------------------- #
#  Public API
# --------------------------------------------------------------------------- #
def save_submission(payload, year, month):
    """Insert/replace this OpCo's submission for the period; log the event; save.
    Returns the number of OpCo submissions now held in the master workbook."""
    # ensure a UTC timestamp exists on the payload
    if not payload.get("submittedAt"):
        payload["submittedAt"] = datetime.datetime.utcnow().isoformat() + "Z"

    wb = _load_wb()
    raws = _read_raw(wb)
    opco = payload.get("opco")
    existed = any(o == opco and str(y) == str(year) and str(m) == str(month)
                  for (o, y, m, _p) in raws)
    raws = [(o, y, m, p) for (o, y, m, p) in raws
            if not (o == opco and str(y) == str(year) and str(m) == str(month))]
    raws.append((opco, str(year), str(month), payload))

    # rewrite hidden _raw
    if "_raw" in wb.sheetnames:
        del wb["_raw"]
    raw = wb.create_sheet("_raw")
    raw.append(["opco", "year", "month", "payload_json"])
    for o, y, m, p in raws:
        raw.append([o, y, m, json.dumps(p, ensure_ascii=False)])
    raw.sheet_state = "hidden"

    _rebuild_flat(wb, raws)
    _append_audit(wb, payload, "updated" if existed else "created")
    _save_wb(wb)
    return len(raws)


def load_submissions():
    wb = _load_wb()
    return [p for (_o, _y, _m, p) in _read_raw(wb) if p]


def master_bytes():
    wb = _load_wb()
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def storage_label():
    """Human-readable note of where the master lives (for the UI)."""
    c = _sp_conf()
    if c:
        return "SharePoint · %s" % c.get("file_path", "submissions.xlsx")
    return "local file · submissions.xlsx"


def list_meta():
    wb = _load_wb()
    out = []
    for opco, year, month, p in _read_raw(wb):
        out.append({"opco": opco, "year": year, "month": month,
                    "submitted_by": p.get("submittedBy", ""),
                    "items_updated": p.get("itemsUpdated", ""),
                    "submitted_at": p.get("submittedAt", "")})
    return out
